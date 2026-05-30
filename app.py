import os
import random
from datetime import datetime
from functools import wraps
from pathlib import Path

from flask import Flask, render_template, request, redirect, url_for, session
from openpyxl import load_workbook
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash

app = Flask(__name__)
app.secret_key = "Z0p0rt3*2026*!!"

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
INSTANCE_DIR = os.path.join(BASE_DIR, "instance")
os.makedirs(INSTANCE_DIR, exist_ok=True)
DB_PATH = os.path.join(INSTANCE_DIR, "quiz.db")
CUESTIONARIOS_DIR = os.path.join(BASE_DIR, "cuestionarios")

app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{DB_PATH}"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(20), nullable=False, default="user")

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated_function

def superadmin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get("role") != "superadmin":
            return redirect(url_for("superadmin_login"))
        return f(*args, **kwargs)
    return decorated_function

HISTORIAL = {}

def listar_cuestionarios():
    carpeta = Path(CUESTIONARIOS_DIR)
    carpeta.mkdir(exist_ok=True)
    return sorted([f.name for f in carpeta.glob("*.xlsx")])

def limpiar_texto(valor):
    if valor is None:
        return ""
    return " ".join(str(valor).strip().split())

def detectar_hoja_valida(wb):
    for nombre in wb.sheetnames:
        ws = wb[nombre]
        if ws.max_row >= 4:
            return nombre
    return wb.sheetnames[0]

def cargar_preguntas_desde_excel(nombre_archivo):
    ruta = os.path.join(CUESTIONARIOS_DIR, nombre_archivo)
    wb = load_workbook(ruta, data_only=True)
    hoja = detectar_hoja_valida(wb)
    ws = wb[hoja]
    preguntas = []
    header_row = 3
    for r in range(header_row + 1, ws.max_row + 1):
        num = ws.cell(row=r, column=1).value
        texto = ws.cell(row=r, column=2).value
        if num is None or texto is None:
            continue
        op_a = ws.cell(row=r, column=3).value
        op_b = ws.cell(row=r, column=4).value
        op_c = ws.cell(row=r, column=5).value
        op_d = ws.cell(row=r, column=6).value
        correcta_excel = ws.cell(row=r, column=7).value
        biblio = ws.cell(row=r, column=8).value
        correcta = None
        if correcta_excel is not None:
            correcta_excel = str(correcta_excel).strip().upper()
            if correcta_excel in ("A", "B", "C", "D"):
                correcta = correcta_excel
        if correcta is None:
            continue
        preguntas.append({
            "numero": int(num),
            "texto": limpiar_texto(texto),
            "A": limpiar_texto(op_a),
            "B": limpiar_texto(op_b),
            "C": limpiar_texto(op_c),
            "D": limpiar_texto(op_d),
            "correcta": correcta,
            "biblio": limpiar_texto(biblio),
        })
    preguntas.sort(key=lambda x: x["numero"])
    return preguntas

@app.route("/register", methods=["GET", "POST"])
def register():
    return redirect(url_for("superadmin_login"))

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip().lower()
        password = request.form.get("password", "")
        user = User.query.filter_by(username=username, role="user").first()
        if user and user.check_password(password):
            session["user_id"] = user.id
            session["usuario"] = user.username
            session["role"] = "user"
            HISTORIAL.setdefault(user.username, [])
            return redirect(url_for("index"))
        return render_template("login.html", error="Usuario o contraseña incorrectos")
    return render_template("login.html")

@app.route("/superadmin/login", methods=["GET", "POST"])
def superadmin_login():
    if request.method == "POST":
        username = request.form.get("username", "").strip().lower()
        password = request.form.get("password", "")
        user = User.query.filter_by(username=username, role="superadmin").first()
        if user and user.check_password(password):
            session["user_id"] = user.id
            session["usuario"] = user.username
            session["role"] = "superadmin"
            return redirect(url_for("superadmin_panel"))
        return render_template("superadmin_login.html", error="Credenciales inválidas")
    return render_template("superadmin_login.html")

@app.route("/superadmin/delete/<int:user_id>", methods=["POST"])
@login_required
@superadmin_required
def superadmin_delete(user_id):
    user = User.query.get_or_404(user_id)
    if user.username == session.get("usuario"):
        return redirect(url_for("superadmin_panel"))
    db.session.delete(user)
    db.session.commit()
    return redirect(url_for("superadmin_panel"))

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/superadmin/panel")
@login_required
@superadmin_required
def superadmin_panel():
    users = User.query.order_by(User.id.desc()).all()
    return render_template("superadmin_panel.html", usuario=session.get("usuario"), users=users)

@app.route("/superadmin/register", methods=["GET", "POST"])
@login_required
@superadmin_required
def superadmin_register():
    if request.method == "POST":
        username = request.form.get("username", "").strip().lower()
        password = request.form.get("password", "")
        role = request.form.get("role", "user")
        if not username or not password:
            return render_template("superadmin_register.html", error="Completa todos los campos", usuario=session.get("usuario"))
        if User.query.filter_by(username=username).first():
            return render_template("superadmin_register.html", error="Ese usuario ya existe", usuario=session.get("usuario"))
        user = User(username=username, role=role)
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        return render_template("superadmin_register.html", success="Usuario creado correctamente", usuario=session.get("usuario"))
    return render_template("superadmin_register.html", usuario=session.get("usuario"))

@app.route("/", methods=["GET", "POST"])
@login_required
def index():
    usuario = session["usuario"]
    historial_usuario = HISTORIAL.get(usuario, [])
    cuestionarios = listar_cuestionarios()
    if request.method == "POST":
        archivo = request.form.get("archivo", "").strip()
        modo = request.form.get("modo", "aleatorio")
        if archivo not in cuestionarios:
            return render_template("index.html", usuario=usuario, historial=historial_usuario, cuestionarios=cuestionarios, error="Selecciona un cuestionario válido")
        preguntas = cargar_preguntas_desde_excel(archivo)
        if not preguntas:
            return render_template("index.html", usuario=usuario, historial=historial_usuario, cuestionarios=cuestionarios, error="Ese archivo no contiene preguntas válidas")
        if modo == "aleatorio":
            try:
                cantidad = int(request.form.get("cantidad", "10"))
            except ValueError:
                cantidad = 10
            if cantidad <= 0:
                cantidad = 10
            cantidad = min(cantidad, len(preguntas))
            seleccion = random.sample(preguntas, cantidad)
        elif modo == "rango":
            try:
                desde = int(request.form.get("desde", "1"))
                hasta = int(request.form.get("hasta", "1"))
            except ValueError:
                return render_template("index.html", usuario=usuario, historial=historial_usuario, cuestionarios=cuestionarios, error="El rango debe ser numérico")
            if desde > hasta:
                desde, hasta = hasta, desde
            seleccion = [p for p in preguntas if desde <= p["numero"] <= hasta]
            if not seleccion:
                return render_template("index.html", usuario=usuario, historial=historial_usuario, cuestionarios=cuestionarios, error="No hay preguntas en ese rango")
        else:
            return render_template("index.html", usuario=usuario, historial=historial_usuario, cuestionarios=cuestionarios, error="Modo no válido")
        return render_template("quiz.html", preguntas=seleccion, usuario=usuario, archivo=archivo, modo=modo)
    return render_template("index.html", usuario=usuario, historial=historial_usuario, cuestionarios=cuestionarios)

@app.route("/corregir", methods=["POST"])
@login_required
def corregir():
    usuario = session["usuario"]
    archivo = request.form.get("archivo", "").strip()
    cuestionarios = listar_cuestionarios()
    if archivo not in cuestionarios:
        return redirect(url_for("index"))
    preguntas = cargar_preguntas_desde_excel(archivo)
    mapa = {p["numero"]: p for p in preguntas}
    respuestas_usuario = {}
    for key, value in request.form.items():
        if key.startswith("q_"):
            num = int(key.split("_", 1)[1])
            respuestas_usuario[num] = value
    total = len(respuestas_usuario)
    correctas = 0
    detalle = []
    for num, resp in respuestas_usuario.items():
        p = mapa.get(num)
        if not p:
            continue
        es_correcta = resp == p["correcta"]
        if es_correcta:
            correctas += 1
        detalle.append({
            "numero": num,
            "texto": p["texto"],
            "marcada": resp,
            "texto_marcada": p.get(resp, ""),
            "correcta": p["correcta"],
            "texto_correcta": p.get(p["correcta"], ""),
            "biblio": p["biblio"],
            "es_correcta": es_correcta,
        })
    detalle.sort(key=lambda x: x["numero"])
    porcentaje = (correctas * 100.0 / total) if total > 0 else 0.0
    intento = {
        "fecha": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "archivo": archivo,
        "total": total,
        "correctas": correctas,
        "porcentaje": porcentaje,
    }
    HISTORIAL.setdefault(usuario, []).append(intento)
    return render_template("resultado.html", total=total, correctas=correctas, porcentaje=porcentaje, detalle=detalle, usuario=usuario, archivo=archivo)

if __name__ == "__main__":
    with app.app_context():
        db.create_all()
    app.run(debug=True, host="0.0.0.0", port=5000)
